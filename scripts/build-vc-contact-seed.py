from __future__ import annotations

import csv
import io
import json
import math
import re
from pathlib import Path
from urllib.parse import urlparse
from typing import Any, NamedTuple, TypeVar

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
CONTACT_INPUT_FILE = Path(r"C:\Users\sreek\Downloads\2000 Plus VC Contact Emails.xlsx")
OPENVC_INPUT_FILE = Path(r"C:\Users\sreek\Downloads\Oct 2025 - OpenVC.xlsx")
OUTPUT_FILE = ROOT / "supabase" / "seed.sql"
TS_OUTPUT_FILE = ROOT / "src" / "data" / "vc-contact-seed.ts"
OPENVC_MIGRATION_FILE = ROOT / "supabase" / "migrations" / "20260531173702_load_openvc_contacts.sql"
T = TypeVar("T")

COORDINATES: dict[str, tuple[float, float]] = {
    "alexandria": (38.8048, -77.0469),
    "amsterdam": (52.3676, 4.9041),
    "atlanta": (33.7490, -84.3880),
    "austin": (30.2672, -97.7431),
    "bangalore": (12.9716, 77.5946),
    "bengaluru": (12.9716, 77.5946),
    "beijing": (39.9042, 116.4074),
    "berlin": (52.5200, 13.4050),
    "boston": (42.3601, -71.0589),
    "boulder": (40.0150, -105.2705),
    "brooklyn": (40.6782, -73.9442),
    "buffalo": (42.8864, -78.8784),
    "burlingame": (37.5779, -122.3481),
    "cambridge": (42.3736, -71.1097),
    "campbell": (37.2872, -121.9500),
    "cayman islands": (19.3133, -81.2546),
    "chicago": (41.8781, -87.6298),
    "cincinnati": (39.1031, -84.5120),
    "copenhagen": (55.6761, 12.5683),
    "dallas": (32.7767, -96.7970),
    "delhi": (28.6139, 77.2090),
    "denver": (39.7392, -104.9903),
    "dubai": (25.2048, 55.2708),
    "durham": (35.9940, -78.8986),
    "greenwich": (41.0262, -73.6282),
    "herzliya": (32.1624, 34.8447),
    "hong kong": (22.3193, 114.1694),
    "houston": (29.7604, -95.3698),
    "hyderabad": (17.3850, 78.4867),
    "jacksonville": (30.3322, -81.6557),
    "london": (51.5072, -0.1276),
    "los altos": (37.3852, -122.1141),
    "los angeles": (34.0522, -118.2437),
    "luxembourg": (49.6116, 6.1319),
    "maryland": (39.0458, -76.6413),
    "menlo park": (37.4530, -122.1817),
    "miami": (25.7617, -80.1918),
    "minneapolis": (44.9778, -93.2650),
    "montreal": (45.5019, -73.5674),
    "mountain view": (37.3861, -122.0839),
    "mumbai": (19.0760, 72.8777),
    "munich": (48.1351, 11.5820),
    "nashville": (36.1627, -86.7816),
    "new york": (40.7128, -74.0060),
    "nyc": (40.7128, -74.0060),
    "oakland": (37.8044, -122.2712),
    "palo alto": (37.4419, -122.1430),
    "paris": (48.8566, 2.3522),
    "portland": (45.5152, -122.6784),
    "portola valley": (37.3841, -122.2352),
    "pune": (18.5204, 73.8567),
    "raleigh": (35.7796, -78.6382),
    "redwood city": (37.4852, -122.2364),
    "salt lake city": (40.7608, -111.8910),
    "san diego": (32.7157, -117.1611),
    "san francisco": (37.7749, -122.4194),
    "san jose": (37.3382, -121.8863),
    "san mateo": (37.5630, -122.3255),
    "santa monica": (34.0195, -118.4912),
    "sausalito": (37.8591, -122.4853),
    "seattle": (47.6062, -122.3321),
    "shanghai": (31.2304, 121.4737),
    "singapore": (1.3521, 103.8198),
    "st louis": (38.6270, -90.1994),
    "stockholm": (59.3293, 18.0686),
    "sydney": (-33.8688, 151.2093),
    "sao paulo": (-23.5558, -46.6396),
    "são paulo": (-23.5558, -46.6396),
    "tel aviv": (32.0853, 34.7818),
    "tokyo": (35.6762, 139.6503),
    "toronto": (43.6532, -79.3832),
    "vancouver": (49.2827, -123.1207),
    "vienna": (48.2082, 16.3738),
    "virginia": (37.4316, -78.6569),
    "washington": (38.9072, -77.0369),
    "washington dc": (38.9072, -77.0369),
    "zurich": (47.3769, 8.5417),
    # Additional OpenVC HQ aliases. Country-only rows use a major startup hub
    # or geographic center so they can still appear in the VC heat map.
    "albuquerque": (35.0844, -106.6504),
    "antwerp": (51.2194, 4.4025),
    "brussels": (50.8503, 4.3517),
    "bucharest": (44.4268, 26.1025),
    "buenos aires": (-34.6037, -58.3816),
    "canada": (43.6532, -79.3832),
    "charlotte": (35.2271, -80.8431),
    "delaware": (39.1582, -75.5244),
    "dublin": (53.3498, -6.2603),
    "france": (48.8566, 2.3522),
    "germany": (52.5200, 13.4050),
    "hamburg": (53.5511, 9.9937),
    "helsinki": (60.1699, 24.9384),
    "india": (28.6139, 77.2090),
    "istanbul": (41.0082, 28.9784),
    "italy": (41.9028, 12.4964),
    "jakarta": (-6.2088, 106.8456),
    "jerusalem": (31.7683, 35.2137),
    "lexington": (38.0406, -84.5037),
    "lisbon": (38.7223, -9.1393),
    "madrid": (40.4168, -3.7038),
    "manchester": (53.4808, -2.2426),
    "melbourne": (-37.8136, 144.9631),
    "mexico city": (19.4326, -99.1332),
    "milan": (45.4642, 9.1900),
    "milano": (45.4642, 9.1900),
    "orange county": (33.7175, -117.8311),
    "orlando": (28.5383, -81.3792),
    "poland": (52.2297, 21.0122),
    "prague": (50.0755, 14.4378),
    "riga": (56.9496, 24.1052),
    "seoul": (37.5665, 126.9780),
    "stuttgart": (48.7758, 9.1829),
    "tallinn": (59.4370, 24.7536),
    "tampa": (27.9506, -82.4572),
    "the hague": (52.0705, 4.3007),
    "uae": (25.2048, 55.2708),
    "uk": (51.5072, -0.1276),
    "united states": (39.8283, -98.5795),
    "us": (39.8283, -98.5795),
    "usa": (39.8283, -98.5795),
    "warsaw": (52.2297, 21.0122),
}


class SourceSpec(NamedTuple):
    path: Path
    label: str
    kind: str


SOURCES: tuple[SourceSpec, ...] = (
    SourceSpec(CONTACT_INPUT_FILE, "2000 Plus VC Contact Emails.xlsx", "contact"),
    SourceSpec(OPENVC_INPUT_FILE, "Oct 2025 - OpenVC.xlsx", "openvc"),
)

OPENVC_STAGE_MAP: dict[str, str] = {
    "1. idea or patent": "Pre-Seed",
    "2. prototype": "Pre-Seed",
    "3. early revenue": "Seed",
    "4. scaling": "Series A",
    "5. growth": "Series B, Series C, Growth",
}

OPENVC_FUND_TYPE_MAP: dict[str, str] = {
    "vc": "Venture Fund",
    "angel network": "Angel Network",
    "angel": "Angel",
    "accelerator": "Accelerator",
    "incubator": "Accelerator",
    "corporate vc": "Corporate VC",
    "family office": "Family Office",
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def to_int(value: Any) -> int | None:
    text = clean(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def normalize_city(location: str) -> str:
    text = location.strip()
    if not text:
        return ""
    first = text.split("/")[0].split(",")[0].strip()
    lower = first.lower()
    if lower in {"nyc", "new york city"}:
        return "New York"
    if lower in {"sf"}:
        return "San Francisco"
    return first


def title_city(alias: str) -> str:
    special_cases = {
        "nyc": "New York",
        "sf": "San Francisco",
        "st louis": "St Louis",
        "sao paulo": "Sao Paulo",
        "sÃ£o paulo": "Sao Paulo",
    }
    return special_cases.get(alias, alias.title())


def coordinate_match(location: str) -> tuple[str, float | None, float | None]:
    city = normalize_city(location)
    key = city.lower()
    if key in COORDINATES:
        latitude, longitude = COORDINATES[key]
        return title_city(key), latitude, longitude

    lower_location = location.lower()
    for alias, coordinates in COORDINATES.items():
        if re.search(rf"(?<![a-z]){re.escape(alias)}(?![a-z])", lower_location):
            latitude, longitude = coordinates
            return title_city(alias), latitude, longitude

    return city, None, None


def coordinates_for(location: str) -> tuple[float | None, float | None]:
    _, latitude, longitude = coordinate_match(location)
    return latitude, longitude


def normalize_name(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def website_domain(value: str) -> str:
    text = value.strip()
    if not text:
        return ""
    if not re.match(r"^[a-z][a-z0-9+.-]*://", text, flags=re.IGNORECASE):
        text = "https://" + text
    try:
        host = urlparse(text).hostname or ""
    except ValueError:
        return ""
    return host.removeprefix("www.").lower()


def sql_literal(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, (int, float)):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def cell(row: tuple[Any, ...], index: dict[str, int], header: str) -> Any:
    position = index[header]
    return row[position] if position < len(row) else None


def chunked(values: list[T], size: int) -> list[list[T]]:
    return [values[index : index + size] for index in range(0, len(values), size)]


def read_rows(source: SourceSpec) -> tuple[list[str], list[tuple[int, tuple[Any, ...]]]]:
    workbook = openpyxl.load_workbook(source.path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    data_rows: list[tuple[int, tuple[Any, ...]]] = []
    for row_number, row in enumerate(rows, start=2):
        if not row or not any(row):
            continue
        data_rows.append((row_number, row))
    return headers, data_rows


def contact_record(
    global_row_number: int,
    source_row_number: int,
    row: tuple[Any, ...],
    index: dict[str, int],
    source_label: str,
) -> tuple[Any, ...]:
    location = clean(cell(row, index, "Location"))
    normalized_city, latitude, longitude = coordinate_match(location)
    investments = to_int(cell(row, index, "Number of Investments")) or 0
    exits = to_int(cell(row, index, "Number of Exits")) or 0

    return (
        global_row_number,
        source_label,
        source_row_number,
        clean(cell(row, index, "Investor Name")),
        clean(cell(row, index, "Fund Type")),
        clean(cell(row, index, "Fund Stage")),
        clean(cell(row, index, "Website")),
        clean(cell(row, index, "Fund Focus (Sectors)")),
        clean(cell(row, index, "Partner Name")),
        clean(cell(row, index, "Partner Email")),
        clean(cell(row, index, "Portfolio Companies")),
        location,
        normalized_city,
        latitude,
        longitude,
        clean(cell(row, index, "Twitter Link")),
        clean(cell(row, index, "LinkedIn Link")),
        clean(cell(row, index, "Facebook Link")),
        investments,
        exits,
        clean(cell(row, index, "Fund Description")),
        to_int(cell(row, index, "Founding Year")),
    )


def normalize_openvc_stage(value: str) -> str:
    stages: list[str] = []
    for raw_stage in value.split(","):
        stage = raw_stage.strip()
        normalized = OPENVC_STAGE_MAP.get(stage.lower(), stage)
        if normalized and normalized not in stages:
            stages.append(normalized)
    return ", ".join(stages)


def normalize_openvc_fund_type(value: str) -> str:
    text = value.strip()
    return OPENVC_FUND_TYPE_MAP.get(text.lower(), text)


def money(value: Any) -> str:
    amount = to_int(value)
    if amount is None:
        return ""
    return f"${amount:,}"


def openvc_record(
    global_row_number: int,
    source_row_number: int,
    row: tuple[Any, ...],
    index: dict[str, int],
    source_label: str,
) -> tuple[Any, ...]:
    location = clean(cell(row, index, "Global HQ"))
    normalized_city, latitude, longitude = coordinate_match(location)
    thesis = clean(cell(row, index, "Investment thesis"))
    countries = clean(cell(row, index, "Countries of investment"))
    first_check_min = money(cell(row, index, "First cheque minimum"))
    first_check_max = money(cell(row, index, "First cheque maximum"))
    check_range = " - ".join(value for value in (first_check_min, first_check_max) if value)
    description_parts = [thesis]
    if countries:
        description_parts.append(f"Invests in: {countries}")
    if check_range:
        description_parts.append(f"First cheque: {check_range}")

    return (
        global_row_number,
        source_label,
        source_row_number,
        clean(cell(row, index, "Investor name")),
        normalize_openvc_fund_type(clean(cell(row, index, "Investor type"))),
        normalize_openvc_stage(clean(cell(row, index, "Stage of investment"))),
        clean(cell(row, index, "Website")),
        thesis,
        "",
        "",
        "",
        location,
        normalized_city,
        latitude,
        longitude,
        "",
        "",
        "",
        0,
        0,
        " ".join(part for part in description_parts if part),
        None,
    )


def seed_record(record: tuple[Any, ...]) -> dict[str, Any]:
    return {
        "id": f"seed-vc-contact-{record[0]}",
        "investorName": record[3],
        "fundType": record[4],
        "fundStage": record[5],
        "website": record[6],
        "fundFocusSectors": record[7],
        "partnerName": record[8],
        "partnerEmail": record[9],
        "portfolioCompanies": record[10],
        "location": record[11],
        "normalizedCity": record[12],
        "latitude": record[13],
        "longitude": record[14],
        "twitterUrl": record[15],
        "linkedinUrl": record[16],
        "facebookUrl": record[17],
        "numberOfInvestments": record[18],
        "numberOfExits": record[19],
        "fundDescription": record[20],
        "foundingYear": record[21],
    }


def records_from_sources() -> tuple[list[tuple[Any, ...]], list[tuple[Any, ...]], int]:
    records: list[tuple[Any, ...]] = []
    openvc_records: list[tuple[Any, ...]] = []
    seen_domains: set[str] = set()
    seen_names: set[str] = set()
    skipped_duplicates = 0

    for source in SOURCES:
        headers, rows = read_rows(source)
        index = {header: headers.index(header) for header in headers}
        for source_row_number, row in rows:
            global_row_number = len(records) + 2
            if source.kind == "contact":
                record = contact_record(global_row_number, source_row_number, row, index, source.label)
            else:
                record = openvc_record(global_row_number, source_row_number, row, index, source.label)

            name_key = normalize_name(record[3])
            domain_key = website_domain(record[6])
            duplicate = bool(domain_key and domain_key in seen_domains) or bool(name_key and name_key in seen_names)
            if source.kind != "contact" and duplicate:
                skipped_duplicates += 1
                continue

            records.append(record)
            if source.kind == "openvc":
                openvc_records.append(record)
            if domain_key:
                seen_domains.add(domain_key)
            if name_key:
                seen_names.add(name_key)

    return records, openvc_records, skipped_duplicates


def write_sql(records: list[tuple[Any, ...]], output_file: Path, include_relax_email_schema: bool = False) -> None:
    if not records:
        return

    columns = [
        "source_row_number",
        "import_source",
        "investor_name",
        "fund_type",
        "fund_stage",
        "website",
        "fund_focus_sectors",
        "partner_name",
        "partner_email",
        "portfolio_companies",
        "location",
        "normalized_city",
        "latitude",
        "longitude",
        "twitter_url",
        "linkedin_url",
        "facebook_url",
        "number_of_investments",
        "number_of_exits",
        "fund_description",
        "founding_year",
    ]
    assignments = ",\n    ".join(f"{column} = excluded.{column}" for column in columns[1:])

    output = io.StringIO()
    output.write("-- Generated from C:\\\\Users\\\\sreek\\\\Downloads\\\\2000 Plus VC Contact Emails.xlsx and C:\\\\Users\\\\sreek\\\\Downloads\\\\Oct 2025 - OpenVC.xlsx\n")
    output.write("-- Regenerate with: python scripts/build-vc-contact-seed.py\n\n")
    if include_relax_email_schema:
        output.write("alter table public.vc_contacts alter column partner_email drop not null;\n")
        output.write("alter table public.vc_contacts drop constraint if exists vc_contacts_partner_email_present;\n\n")

    for batch in chunked(records, 250):
        output.write(f"insert into public.vc_contacts ({', '.join(columns)})\nvalues\n")
        output.write(
            ",\n".join(
                "  (" + ", ".join(sql_literal(value) for value in (record[0], record[1], *record[3:])) + ")"
                for record in batch
            )
        )
        output.write(
            "\non conflict (source_row_number) do update set\n"
            f"    {assignments},\n"
            "    updated_at = now();\n\n"
        )

    output_file.write_text(output.getvalue(), encoding="utf-8", newline="\n")


def main() -> None:
    records, openvc_records, skipped_duplicates = records_from_sources()
    seed_records = [seed_record(record) for record in records]

    write_sql(records, OUTPUT_FILE)
    write_sql(openvc_records, OPENVC_MIGRATION_FILE, include_relax_email_schema=True)
    TS_OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    TS_OUTPUT_FILE.write_text(
        "import type { VCContact } from '@/lib/apparent-types';\n\n"
        "// Generated from C:\\\\Users\\\\sreek\\\\Downloads\\\\2000 Plus VC Contact Emails.xlsx and C:\\\\Users\\\\sreek\\\\Downloads\\\\Oct 2025 - OpenVC.xlsx\n"
        "// Regenerate with: python scripts/build-vc-contact-seed.py\n"
        f"export const vcContactSeed = {json.dumps(seed_records, ensure_ascii=False, separators=(',', ':'))} satisfies VCContact[];\n",
        encoding="utf-8",
        newline="\n",
    )
    plotted = sum(1 for record in records if record[13] is not None and record[14] is not None)
    openvc_plotted = sum(1 for record in openvc_records if record[13] is not None and record[14] is not None)
    print(f"Wrote {len(records)} VC contacts to {OUTPUT_FILE}")
    print(f"Wrote local app fallback to {TS_OUTPUT_FILE}")
    print(f"Wrote {len(openvc_records)} new OpenVC contacts to {OPENVC_MIGRATION_FILE}")
    print(f"Skipped {skipped_duplicates} duplicate OpenVC contacts")
    print(f"{plotted} contacts include derived coordinates for heat-map plotting")
    print(f"{openvc_plotted} OpenVC contacts include derived coordinates for heat-map plotting")


if __name__ == "__main__":
    main()
